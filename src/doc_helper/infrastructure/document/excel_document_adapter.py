"""Excel document adapter using openpyxl.

Note: Using openpyxl instead of xlwings for cross-platform compatibility.
openpyxl works on Linux/Mac/Windows without requiring Excel installation.
"""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from doc_helper.domain.common.result import Failure, Result, Success
from doc_helper.domain.document.document_adapter import IDocumentAdapter
from doc_helper.domain.document.document_format import DocumentFormat


class ExcelDocumentAdapter(IDocumentAdapter):
    """Excel document adapter using openpyxl.

    Generates .xlsx documents by replacing named ranges or cell markers
    in templates.

    Cells in Excel templates should contain field markers like {{field_name}}.
    For example, a cell containing "{{project_name}}" will be replaced
    with the value of field_values["project_name"].

    Example:
        adapter = ExcelDocumentAdapter()
        result = adapter.generate(
            template_path="template.xlsx",
            output_path="output.xlsx",
            field_values={"project_name": "Soil Investigation"}
        )
    """

    @property
    def format(self) -> DocumentFormat:
        return DocumentFormat.EXCEL

    def generate(
        self,
        template_path: str | Path,
        output_path: str | Path,
        field_values: dict[str, Any],
    ) -> Result[None, str]:
        """Generate Excel document from template.

        Args:
            template_path: Path to .xlsx template
            output_path: Path for generated .xlsx document
            field_values: Field values to insert

        Returns:
            Success(None) if generated, Failure(error) otherwise
        """
        # Validate inputs
        if not isinstance(template_path, (str, Path)):
            return Failure("template_path must be a string or Path")
        if not isinstance(output_path, (str, Path)):
            return Failure("output_path must be a string or Path")
        if not isinstance(field_values, dict):
            return Failure("field_values must be a dict")

        template_path = Path(template_path)
        output_path = Path(output_path)

        # Validate template
        validation = self.validate_template(template_path)
        if isinstance(validation, Failure):
            return validation

        try:
            # Load template
            workbook = load_workbook(str(template_path))

            # Replace markers in all worksheets
            for sheet in workbook.worksheets:
                self._replace_markers_in_sheet(sheet, field_values)

            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            workbook.save(str(output_path))

            return Success(None)

        except Exception as e:
            return Failure(f"Error generating Excel document: {str(e)}")

    def validate_template(self, template_path: str | Path) -> Result[None, str]:
        """Validate Excel template.

        Args:
            template_path: Path to template file

        Returns:
            Success(None) if valid, Failure(error) otherwise
        """
        if not isinstance(template_path, (str, Path)):
            return Failure("template_path must be a string or Path")

        template_path = Path(template_path)

        if not template_path.exists():
            return Failure(f"Template not found: {template_path}")

        if not template_path.is_file():
            return Failure(f"Template path is not a file: {template_path}")

        if template_path.suffix.lower() not in [".xlsx", ".xlsm"]:
            return Failure(
                f"Template must be .xlsx or .xlsm file: {template_path}"
            )

        # Try to load workbook
        try:
            load_workbook(str(template_path))
            return Success(None)
        except Exception as e:
            return Failure(f"Invalid Excel document: {str(e)}")

    def _replace_markers_in_sheet(
        self, sheet: Any, field_values: dict[str, Any]
    ) -> None:
        """Replace field markers in worksheet.

        Args:
            sheet: Worksheet object
            field_values: Field values to insert
        """
        # Iterate through all cells
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # Check if cell contains markers
                if isinstance(cell.value, str) and "{{" in cell.value:
                    # Replace all markers in cell
                    new_value = cell.value
                    for field_name, field_value in field_values.items():
                        marker = f"{{{{{field_name}}}}}"
                        if marker in new_value:
                            replacement = (
                                str(field_value) if field_value is not None else ""
                            )
                            new_value = new_value.replace(marker, replacement)

                    cell.value = new_value
