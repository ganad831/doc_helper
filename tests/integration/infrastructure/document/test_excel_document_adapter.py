"""Integration tests for ExcelDocumentAdapter."""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from doc_helper.domain.common.result import Failure, Success
from doc_helper.domain.document.document_format import DocumentFormat
from doc_helper.infrastructure.document.excel_document_adapter import (
    ExcelDocumentAdapter,
)


class TestExcelDocumentAdapter:
    """Integration tests for ExcelDocumentAdapter."""

    @pytest.fixture
    def adapter(self) -> ExcelDocumentAdapter:
        """Create adapter instance."""
        return ExcelDocumentAdapter()

    @pytest.fixture
    def template_file(self, tmp_path: Path) -> Path:
        """Create a simple Excel template with field markers."""
        template_path = tmp_path / "template.xlsx"

        # Create a simple Excel workbook with markers
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Project Name:"
        ws["B1"] = "{{project_name}}"
        ws["A2"] = "Date:"
        ws["B2"] = "{{date}}"
        ws["A3"] = "Description:"
        ws["B3"] = "{{description}}"

        wb.save(str(template_path))
        return template_path

    def test_adapter_format(self, adapter: ExcelDocumentAdapter) -> None:
        """Adapter should report EXCEL format."""
        assert adapter.format == DocumentFormat.EXCEL

    def test_validate_template_success(
        self, adapter: ExcelDocumentAdapter, template_file: Path
    ) -> None:
        """validate_template should succeed for valid template."""
        result = adapter.validate_template(template_file)
        assert isinstance(result, Success)

    def test_validate_template_not_found(
        self, adapter: ExcelDocumentAdapter
    ) -> None:
        """validate_template should fail for non-existent template."""
        result = adapter.validate_template("nonexistent.xlsx")
        assert isinstance(result, Failure)
        assert "not found" in result.error.lower()

    def test_validate_template_invalid_file(
        self, adapter: ExcelDocumentAdapter, tmp_path: Path
    ) -> None:
        """validate_template should fail for invalid Excel file."""
        # Create a non-Excel file
        invalid_file = tmp_path / "invalid.xlsx"
        invalid_file.write_text("Not an Excel document")

        result = adapter.validate_template(invalid_file)
        assert isinstance(result, Failure)
        assert "invalid" in result.error.lower()

    def test_generate_document_success(
        self, adapter: ExcelDocumentAdapter, template_file: Path, tmp_path: Path
    ) -> None:
        """generate should create output document."""
        output_path = tmp_path / "output.xlsx"
        field_values = {
            "project_name": "Test Project",
            "date": "2024-01-15",
            "description": "Test Description",
        }

        result = adapter.generate(
            template_path=template_file,
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Success)
        assert output_path.exists()

    def test_generate_replaces_markers(
        self, adapter: ExcelDocumentAdapter, template_file: Path, tmp_path: Path
    ) -> None:
        """generate should replace field markers with values."""
        output_path = tmp_path / "output.xlsx"
        field_values = {
            "project_name": "My Project",
            "date": "2024-01-15",
            "description": "My Description",
        }

        result = adapter.generate(
            template_path=template_file,
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Success)

        # Verify markers were replaced
        wb = load_workbook(str(output_path))
        ws = wb.active
        assert ws["B1"].value == "My Project"
        assert ws["B2"].value == "2024-01-15"
        assert ws["B3"].value == "My Description"

    def test_generate_handles_none_values(
        self, adapter: ExcelDocumentAdapter, template_file: Path, tmp_path: Path
    ) -> None:
        """generate should handle None values as empty cells."""
        output_path = tmp_path / "output.xlsx"
        field_values = {
            "project_name": "Test",
            "date": None,
            "description": "Desc",
        }

        result = adapter.generate(
            template_path=template_file,
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Success)

        wb = load_workbook(str(output_path))
        ws = wb.active
        # When marker is replaced with empty string, Excel treats it as None (empty cell)
        assert ws["B2"].value is None or ws["B2"].value == ""

    def test_generate_preserves_unmarked_cells(
        self, adapter: ExcelDocumentAdapter, template_file: Path, tmp_path: Path
    ) -> None:
        """generate should preserve cells without markers."""
        output_path = tmp_path / "output.xlsx"
        field_values = {"project_name": "Test"}

        result = adapter.generate(
            template_path=template_file,
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Success)

        wb = load_workbook(str(output_path))
        ws = wb.active
        assert ws["A1"].value == "Project Name:"  # Preserved
        assert ws["A2"].value == "Date:"  # Preserved

    def test_generate_with_multiple_sheets(
        self, adapter: ExcelDocumentAdapter, tmp_path: Path
    ) -> None:
        """generate should process all sheets in workbook."""
        template_path = tmp_path / "template.xlsx"

        # Create workbook with multiple sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "{{field1}}"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "{{field2}}"

        wb.save(str(template_path))

        output_path = tmp_path / "output.xlsx"
        field_values = {"field1": "Value1", "field2": "Value2"}

        result = adapter.generate(
            template_path=template_path,
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Success)

        wb = load_workbook(str(output_path))
        assert wb["Sheet1"]["A1"].value == "Value1"
        assert wb["Sheet2"]["A1"].value == "Value2"

    def test_generate_creates_parent_directories(
        self, adapter: ExcelDocumentAdapter, template_file: Path, tmp_path: Path
    ) -> None:
        """generate should create parent directories if needed."""
        output_path = tmp_path / "subdir" / "output.xlsx"
        field_values = {}

        result = adapter.generate(
            template_path=template_file,
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Success)
        assert output_path.exists()
        assert output_path.parent.exists()

    def test_generate_with_invalid_template(
        self, adapter: ExcelDocumentAdapter, tmp_path: Path
    ) -> None:
        """generate should fail with invalid template."""
        output_path = tmp_path / "output.xlsx"
        field_values = {}

        result = adapter.generate(
            template_path="nonexistent.xlsx",
            output_path=output_path,
            field_values=field_values,
        )

        assert isinstance(result, Failure)
